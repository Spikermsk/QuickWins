#!/usr/bin/env python

from netmiko import ConnectHandler
from getpass import getpass
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.table import Table


def main():

    username, password = get_creds()

    primary = {
        'device_type': 'cisco_asa',
        'host': 'fwl-dc1-vpn-a',
        'username': username,
        'password': password,
    }

    secondary = {
        'device_type': 'cisco_asa',
        'host': 'fwl-dc0-inet-a',
        'username': username,
        'password': password,
    }

    now = datetime.now()
    tab_name = now.strftime('%Y_%m_%d_%H_%M_%S')

    results = []
    for firewall in [primary, secondary]:
        results.append(firewall['host'])
        vpn_sessiondb = show_vpn_sessiondb(firewall)
        results.append(vpn_sessiondb)

    output_to_excel(tab_name, results)


def get_creds():
    '''
    '''

    print('-' * 40)

    un = input('Username, (q) to quit: ')
    if un.lower() == 'q': exit('EXITING')
    pw = getpass()
    if pw.lower() == 'q': exit('EXITING')

    print('-' * 40)

    return(un, pw)


def show_vpn_sessiondb(device):
    '''
    '''

    done = False
    while not done:
        try:
            net_connect = ConnectHandler(**device)
            print('Gathering information from {}'.format(device['host']))
            done = True
        except:
            print('\n')
            print('ERROR: Invalid username or password for {}'.format(device['host']))
            username, password = get_creds()
            device['username'] = username
            device['password'] = password

    output = net_connect.send_command('show vpn-sessiondb anyconnect', use_textfsm=True)
    net_connect.disconnect()
    return(output)


def output_to_excel(tab, data):
    '''
    '''

    PATH = r'S:\Cit\Operations\Network\AnyConnect'
    FILE = r'\AnyConnect.xlsx'
    excel_file = PATH + FILE

    try:
        wb = load_workbook(filename=excel_file)
    except:
        wb = Workbook()
        ws = wb.active
        wb.remove(ws)

    ws = wb.create_sheet(tab, 0)

    ws.cell(1, 1, 'Firewall')
    column_number = {}
    # Using the first dictionary from the second item
    # in the list to generate the column headings
    headings = data[1][0].keys()
    number_of_columns = len(headings) + 1
    max_column = get_column_letter(number_of_columns)
    for element, heading in enumerate(headings, 2):
        # Populating the column header cells
        ws.cell(1, element, heading)
        # Creating a dictionary to store the column headers
        # with their column number because dictionaries are unordered
        column_number[heading] = element

    row_number = 1
    # Iterating through the list
    for item in data:
        if type(item) == str:
            hostname = item
        else:
            # Iterating through the list of dictionaries
            for row_data in item:
                row_number += 1
                # Populating the first cell in the row with the firewall name
                ws.cell(row_number, 1, hostname)
                for column_heading in headings:
                    # Populating the rest of the cells in the row by using
                    # a dictionary lookup where the column header is the key
                    ws.cell(row_number, column_number.get(column_heading), row_data.get(column_heading))

    table_ref = 'A1:{}{}'.format(max_column, row_number)
    table_name = '_{}'.format(tab)
    vpn_table = Table(displayName=table_name, ref=table_ref)
    ws.add_table(vpn_table)
    ws.freeze_panes = 'D2'
    wb.save(excel_file)
    print('Recorded {} rows in spreadsheet {}'.format(row_number, excel_file))


if __name__ == "__main__":
    main()
